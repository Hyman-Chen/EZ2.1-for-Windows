import csv
from pathlib import Path
import os
import sys
import time

from ttkbootstrap.dialogs import Messagebox
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from openpyxl import load_workbook
import pandas as pd

import initTool
import functions

BASE_PATH = os.path.dirname(os.path.realpath(sys.argv[0]))
IMG_PATH = Path(BASE_PATH) / 'data' / 'image'

SAVE_PATH = Path(BASE_PATH) / 'data' / 'save'
tempDATA_PATH = Path(BASE_PATH) / 'data' / 'temp'

appendXlData_PATH = Path(tempDATA_PATH) / 'AD'
appendcsvData_PATH = Path(tempDATA_PATH) / 'ADcsv'

DATA_PATH = Path(BASE_PATH)


class ZongbiaozhizuoFrame(ttk.Frame):
    '''手动导入'''

    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        self.root = master
        self.root.showframe.destroy()
        self.root.showframe = self
        self.pack(side=TOP, fill=BOTH, expand=True)
        # 创建temp文件夹
        if not os.path.exists(Path.joinpath(SAVE_PATH)):
            os.makedirs(Path.joinpath(tempDATA_PATH))
        if not os.path.exists(Path.joinpath(tempDATA_PATH, 'AD')):
            os.makedirs(Path.joinpath(tempDATA_PATH, 'AD'))
        if not os.path.exists(Path.joinpath(tempDATA_PATH, 'ADcsv')):
            os.makedirs(Path.joinpath(tempDATA_PATH, 'ADcsv'))
        # 　设置修改后的路径
        self.DATA_PATH = Path(BASE_PATH)
        if str(pd.read_csv(Path.joinpath(SAVE_PATH, 'setting.csv'))['data_base_path'][0]) != str(
                self.DATA_PATH):
            self.DATA_PATH = Path(pd.read_csv(Path.joinpath(SAVE_PATH, 'setting.csv'))['data_base_path'][0])
        # 图片
        self.images = [
            ttk.PhotoImage(file=IMG_PATH / 'icons8_add_folder_24px.png'),
        ]
        # 初始化主题
        setting_df = pd.read_csv(Path.joinpath(SAVE_PATH, 'setting.csv'))
        theme_name = setting_df['current_theme'][0]
        style = ttk.Style()
        style.theme_use(theme_name)

        self.create_paget()
        self.create_pageb()

    def create_paget(self):
        '''1 paget'''
        self.framet = ttk.LabelFrame(self, text='手动导入', bootstyle=SUCCESS)
        self.framet.pack(side=TOP, fill=BOTH, expand=False, padx=10, pady=10)
        # 1.1 导入名单
        ttk.Button(
            self.framet,
            text='导入名单',
            bootstyle=('outline', WARNING),
            width=9,
            command=functions.importfile,
            image=self.images[0],
            compound=LEFT
        ).pack(side=TOP, fill=X, expand=TRUE, padx=20, pady=10)
        # 1.2 开始导入
        mid_frame = ttk.Frame(self.framet)
        mid_frame.pack(side=TOP, fill=X, expand=TRUE)
        ttk.Button(
            mid_frame,
            text='开始导入',
            bootstyle=SUCCESS,
            width=9,
            padding=(0, 8),
            command=self.append
        ).pack(side=LEFT, fill=X, expand=TRUE, padx=(20, 10), pady=0)
        # 1.3 导出总表
        ttk.Button(
            mid_frame,
            text='导出总表',
            bootstyle=SUCCESS,
            width=9,
            padding=(0, 8),
            command=self.outPut
        ).pack(side=LEFT, fill=X, expand=TRUE, padx=(10, 20), pady=0)
        # 1.4 选中并删除
        ttk.Button(
            self.framet,
            text='选中并删除',
            bootstyle=DANGER,
            width=9,
            padding=(0, 8),
            command=self.dele
        ).pack(side=TOP, fill=X, expand=TRUE, padx=20, pady=10)
        # 1.#
        ttk.Label(
            master=self.framet,
            text='注：先选择需要导入总表的培训名单！',
            width=30,
        ).pack(side=TOP, fill=X, expand=TRUE, padx=20, pady=(0, 10))

    def create_pageb(self):
        '''2 pageb'''
        # 放在frameb上的笔记本Notebook
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        # 名单总表
        self.frame1 = ttk.Frame(self.notebook)
        self.notebook.add(self.frame1, text="名单总表")
        self.tableValues1 = initTool.Files().csv_to_treeviewlist(SAVE_PATH, 'prjData.csv')[1:]
        self.tableColumns1 = ['序号', '单位', '姓名', '工种', '性别', '年龄', '身份证号', '家庭住址', '手机', '培训日期', '备注']
        A = initTool.Tab1_2(self.frame1, self.tableValues1, self.tableColumns1)
        self.treeview1 = A.treeview

    def append(self):
        '''导入总表'''
        df = pd.read_csv(Path.joinpath(SAVE_PATH, 'prjData.csv'))
        try:
            # 判断是否重复导入
            last_date = df.loc[df.shape[0] - 1]['培训日期'].split('.')
            current_date = time.strftime("%Y.%m.%d").split('.')
            if len(last_date) != 3:
                last_date = str(df.loc[df.shape[0] - 1]['培训日期']).split('-')
            for index in range(2):
                last_date[index] = str(last_date[index])
            last_date[2] = str(last_date[2][:2])
            for index in range(len(current_date)):
                current_date[index] = str(current_date[index])
            if current_date == last_date:
                result = Messagebox.show_question1('今天已导入过文件，是否继续导入？')
                if result == '是':
                    self.doit()
                else:
                    pass
            else:
                self.doit()
        except:
            self.doit()

    def doit(self):
        '''将名单导入总表子函数'''
        initTool.Files().del_temp_csv_ADcsv()
        filelist = initTool.Files().get_file_name(appendXlData_PATH)
        for i in filelist:
            myworksheet = \
            load_workbook(Path.joinpath(appendXlData_PATH, str(i[1]) + '.xlsx'), read_only=True).worksheets[
                0]
            initTool.Files().file_to_csv(myworksheet, appendcsvData_PATH, 'file' + str(i[0]) + '.csv')
            df = pd.read_csv(Path.joinpath(appendcsvData_PATH, 'file' + str(i[0]) + '.csv'))
            if myworksheet.cell(row=2, column=11).value != '':
                for row in range(df.shape[0]):
                    date = myworksheet.cell(row=2, column=11).value
                    df.loc[row, '退场时间'] = date
            else:
                pass
            df.to_csv(Path.joinpath(appendcsvData_PATH, 'file' + str(i[0]) + '.csv'), index=False)
        with open(Path.joinpath(SAVE_PATH, 'prjData.csv'), 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            lastRow = pd.read_csv(Path.joinpath(SAVE_PATH, 'prjData.csv')).shape[0]
            for i in filelist:  # 遍历需要添加的文件
                df = pd.read_csv(Path.joinpath(appendcsvData_PATH, ('file' + str(i[0]) + '.csv')))
                for row in range(df.shape[0]):
                    addList = []
                    addList.append(lastRow + 1)
                    for item in df.iloc[row, [m for m in range(1, 11)]]:
                        addList.append(item)
                    writer.writerow(addList)
                    lastRow += 1
            f.close()
        self.fresh(self.treeview1, initTool.Files().csv_to_treeviewlist(SAVE_PATH, 'prjData.csv'))
        Messagebox.show_info_success('导入成功！', '温馨提示')

    def outPut(self):
        '''导出总表'''
        output_name = pd.read_csv(Path.joinpath(SAVE_PATH, 'setting.csv'))['simplified_name'][0]
        if output_name == " ":
            Messagebox.show_info_success('请设置正确格式的项目简称！', '温馨提示')
        initTool.Files().csv_to_excel2(SAVE_PATH, 'prjData.csv', self.DATA_PATH, str(output_name) + '.xlsx')
        Messagebox.show_info_success('导出成功！', '温馨提示')

    #
    def dele(self):
        '''删除表中数据'''
        try:
            a = int(self.treeview1.selection()[0])
            for selection in self.treeview1.selection():
                self.rowval = int(selection)
                pd.read_csv(Path.joinpath(SAVE_PATH, 'prjData.csv')).drop([a - 1]).to_csv(
                    Path.joinpath(SAVE_PATH, 'prjData.csv'), index=False)
                initTool.Files().fresh_csv(SAVE_PATH, 'prjData.csv')
                self.treeview1.delete(selection)
                self.fresh(self.treeview1, initTool.Files().csv_to_treeviewlist(SAVE_PATH, 'prjData.csv'))
            Messagebox.show_info_success('已删除！', '温馨提示')
        except:
            Messagebox.show_info_success('请选择需要删除的人员！', '温馨提示')

    def fresh(self, treeview, tableValues):
        '''刷新页面'''
        for _ in map(treeview.delete, treeview.get_children("")):
            pass
        try:
            row = 1
            for data in tableValues[1:]:
                treeview.insert('', 'end', text=row, value=data, iid=str(row))
                row += 1
        except:
            pass
