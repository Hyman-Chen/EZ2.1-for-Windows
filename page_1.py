from pathlib import Path
import datetime
import os
import sys
import time
import csv

from ttkbootstrap.dialogs import Messagebox
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import pandas as pd
from openpyxl import load_workbook

import initTool
import functions

# 初始化源文件路径
BASE_PATH = os.path.dirname(os.path.realpath(sys.argv[0]))
IMG_PATH = Path(BASE_PATH) / 'data' / 'image'
SAVE_PATH = Path(BASE_PATH) / 'data' / 'save'
tempDATA_PATH = Path(BASE_PATH) / 'data' / 'temp'
csvFilesToBeMerged_PATH = Path(tempDATA_PATH) / 'csvFiles'
appendXlData_PATH = Path(tempDATA_PATH) / 'AD'
appendcsvData_PATH = Path(tempDATA_PATH) / 'ADcsv'

# 输出表格位置
Mer_PATH = Path(BASE_PATH)
DATA_PATH = Path(BASE_PATH)


class MingdanhuizongFrame(ttk.Frame):
    '''名单汇总'''

    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        self.root = master
        self.root.showframe.destroy()
        self.root.showframe = self
        self.pack(side=TOP, fill=BOTH, expand=True)
        # 创建temp文件夹
        if not os.path.exists(Path.joinpath(SAVE_PATH)):
            os.makedirs(Path.joinpath(tempDATA_PATH))
        if not os.path.exists(Path.joinpath(tempDATA_PATH, 'XlFiles')):
            os.makedirs(Path.joinpath(tempDATA_PATH, 'XlFiles'))
        if not os.path.exists(Path.joinpath(tempDATA_PATH, 'csvFiles')):
            os.makedirs(Path.joinpath(tempDATA_PATH, 'csvFiles'))
        # 定义修改后的地址
        self.Mer_PATH = Path(BASE_PATH)
        if str(pd.read_csv(Path.joinpath(SAVE_PATH, 'setting.csv'))['merge_file_path'][0]) != str(
                self.Mer_PATH):
            self.Mer_PATH = Path(pd.read_csv(Path.joinpath(SAVE_PATH, 'setting.csv'))['merge_file_path'][0])
        # 图片
        self.images = [
            ttk.PhotoImage(file=IMG_PATH / 'icons8_folder_24px.png')
        ]
        # 设置主题
        setting_df = pd.read_csv(Path.joinpath(SAVE_PATH, 'setting.csv'))
        theme_name = setting_df['current_theme'][0]
        style = ttk.Style()
        style.theme_use(theme_name)

        self.create_paget()
        self.create_pageb()

    def create_paget(self):
        '''1 paget'''
        self.framet = ttk.LabelFrame(self, text='名单汇总', bootstyle=SUCCESS)
        self.framet.pack(side=TOP, fill=BOTH, expand=False, padx=10, pady=10)
        # 1.1 选择文件
        ttk.Button(
            self.framet,
            text='选择文件',
            bootstyle=('outline', WARNING),
            width=9,
            command=self.select_files,
            image=self.images[0],
            compound=LEFT
        ).pack(side=TOP, fill=X, padx=20, pady=10)
        # 1.2 合并并保存
        midframe = ttk.Frame(self.framet)
        midframe.pack(side=TOP, fill=BOTH, expand=TRUE)
        ttk.Button(
            midframe,
            text='合并并保存',
            bootstyle=('default', SUCCESS),
            width=9,
            padding=(0, 8),
            command=self.merge_files
        ).pack(side=LEFT, fill=X, expand=TRUE, padx=(20, 10))

        # 1.3 从名单中删除
        ttk.Button(
            midframe,
            text='从名单中删除',
            bootstyle=('default', DANGER),
            width=9,
            padding=(0, 8),
            command=self.shanchu
        ).pack(side=LEFT, fill=X, expand=TRUE, padx=(10, 20))

        # 1.#
        ttk.Label(
            master=self.framet,
            text=f'注：合并前先选择文件，合并后的文件保存位置为\n"%s"。' % self.Mer_PATH,
            padding=(20, 0)
        ).pack(side=LEFT, fill=BOTH, expand=TRUE, pady=10)

    def create_pageb(self):
        '''2 pageb'''
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(side=BOTTOM, fill=BOTH, expand=True, padx=10, pady=(0, 10))
        # 2.1 待合并文件预览
        self.frame1 = ttk.Frame(self.notebook)
        self.notebook.add(self.frame1, text="待合并文件预览")
        self.tableValues1 = initTool.Files().get_file_name(path=Path.joinpath(tempDATA_PATH, 'XlFiles'))
        self.tableColumns1 = ['序号', '文件名']
        A = initTool.Tab1_1(self.frame1, self.tableValues1, self.tableColumns1)
        self.treeview1 = A.treeview
        self.treeview1.column('#1', width=10)
        # 2.2 今日名单预览
        self.frame2 = ttk.Frame(self.notebook)
        self.notebook.add(self.frame2, text="今日名单预览")
        self.tableValues2 = []
        self.tableColumns2 = ['序号', '单位', '姓名', '工种', '性别', '年龄', '身份证号', '家庭住址', '手机', '培训日期', '备注']
        B = initTool.Tab1_2(self.frame2, self.tableValues2, self.tableColumns2)
        self.treeview2 = B.treeview
        self.treeview2.column('#2', width=70)

    def select_files(self):
        '''选择文件'''
        functions.select_files(self)

    def merge_files(self):
        '''合并文件'''
        functions.merge_files(self)
        self.append()
    def shanchu(self):
        '''选中并删除'''
        try:
            self.selected_item = self.treeview2.selection()[0]
            self.rowval = int(self.selected_item)
            pd.read_csv(Path.joinpath(csvFilesToBeMerged_PATH, 'total.csv')).drop([self.rowval - 1]).to_csv(
                Path.joinpath(csvFilesToBeMerged_PATH, 'total.csv'), index=False)
            initTool.Files().fresh_csv(csvFilesToBeMerged_PATH, 'total.csv')
            self.treeview2.delete(self.selected_item)
            self.fresh()
            today = datetime.date.today()
            month = int(today.month)
            day = int(today.day)
            save_name = pd.read_csv(Path.joinpath(SAVE_PATH, 'setting.csv'))['simplified_name'][0]
            filename = str(month) + "月" + str(day) + "日" + str(save_name) + "培训名单" + ".xlsx"
            initTool.Files().csv_to_excel(today, self.Mer_PATH, filename)
            Messagebox.show_info_success('保存成功！', '温馨提示')
        except:
            Messagebox.show_info_success('请选择需要删除的人员！', '温馨提示')

    def fresh(self):
        '''刷新页面'''
        for _ in map(self.treeview2.delete, self.treeview2.get_children("")):
            pass
        try:
            self.tableValues2 = initTool.Files().csv_to_treeviewlist(csvFilesToBeMerged_PATH, 'total.csv')
            row = 1
            for data in self.tableValues2[1:]:
                self.treeview2.insert('', 'end', text=row, value=data, iid=str(row))
                row += 1
        except:
            pass

    def append(self):
        '''导入总表'''
        df = pd.read_csv(Path.joinpath(SAVE_PATH, 'prjData.csv'))
        try:
            # 判断是否重复导入
            last_date = df.loc[df.shape[0] - 1]['培训日期'].split('.')
            current_date = time.strftime("%Y.%m.%d").split('.')
            if len(last_date) != 3:
                last_date = str(df.loc[df.shape[0] - 1]['培训日期']).split('-')
            for index in range(2):
                last_date[index] = str(last_date[index])
            last_date[2] = str(last_date[2][:2])
            for index in range(len(current_date)):
                current_date[index] = str(current_date[index])
            if current_date == last_date:
                result = Messagebox.show_question1('今天已导入过文件，是否继续导入？')
                if result == '是':
                    self.doit()
                else:
                    pass
            else:
                self.doit()
        except:
            self.doit()

    def doit(self):
        '''将名单导入总表子函数'''
        initTool.Files().del_temp_csv_ADcsv()
        filelist = initTool.Files().get_file_name(appendXlData_PATH)
        for i in filelist:
            myworksheet = \
            load_workbook(Path.joinpath(appendXlData_PATH, str(i[1]) + '.xlsx'), read_only=True).worksheets[
                0]
            initTool.Files().file_to_csv(myworksheet, appendcsvData_PATH, 'file' + str(i[0]) + '.csv')
            df = pd.read_csv(Path.joinpath(appendcsvData_PATH, 'file' + str(i[0]) + '.csv'))
            if myworksheet.cell(row=2, column=11).value != '':
                for row in range(df.shape[0]):
                    date = myworksheet.cell(row=2, column=11).value
                    df.loc[row, '退场时间'] = date
            else:
                pass
            df.to_csv(Path.joinpath(appendcsvData_PATH, 'file' + str(i[0]) + '.csv'), index=False)
        with open(Path.joinpath(SAVE_PATH, 'prjData.csv'), 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            lastRow = pd.read_csv(Path.joinpath(SAVE_PATH, 'prjData.csv')).shape[0]
            for i in filelist:  # 遍历需要添加的文件
                df = pd.read_csv(Path.joinpath(appendcsvData_PATH, ('file' + str(i[0]) + '.csv')))
                for row in range(df.shape[0]):
                    addList = []
                    addList.append(lastRow + 1)
                    for item in df.iloc[row, [m for m in range(1, 11)]]:
                        addList.append(item)
                    writer.writerow(addList)
                    lastRow += 1
            f.close()
