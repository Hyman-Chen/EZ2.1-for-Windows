import tkinter as tk
from pathlib import Path
import os
import sys

from ttkbootstrap.dialogs import Messagebox
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment

import initTool
import functions

BASE_PATH = os.path.dirname(os.path.realpath(sys.argv[0]))
IMG_PATH = Path(BASE_PATH) / 'data' / 'image'

tempDATA_PATH = Path(BASE_PATH) / 'data' / 'temp'
SAVE_PATH = Path(BASE_PATH) / 'data' / 'save'
kaoqinXlData_PATH = Path(tempDATA_PATH) / 'KQ'
kaoqincsvData_PATH = Path(tempDATA_PATH) / 'KQcsv'

KQ_PATH = Path(BASE_PATH)


class XianchangqingkuangFrame(ttk.Frame):
    '''现场人数'''

    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        self.root = master
        self.root.showframe.destroy()
        self.root.showframe = self
        self.pack(side=TOP, fill=BOTH, expand=True)
        # 创建temp文件夹
        if not os.path.exists(Path.joinpath(SAVE_PATH)):
            os.makedirs(Path.joinpath(tempDATA_PATH))
        if not os.path.exists(Path.joinpath(tempDATA_PATH, 'KQ')):
            os.makedirs(Path.joinpath(tempDATA_PATH, 'KQ'))
        if not os.path.exists(Path.joinpath(tempDATA_PATH, 'KQcsv')):
            os.makedirs(Path.joinpath(tempDATA_PATH, 'KQcsv'))
        # 定义修改后的导出位置
        self.KQ_PATH = Path(BASE_PATH)
        if str(pd.read_csv(Path.joinpath(SAVE_PATH, 'setting.csv'))['kaoqin_path'][0]) != str(
                self.KQ_PATH):
            self.KQ_PATH = Path(pd.read_csv(Path.joinpath(SAVE_PATH, 'setting.csv'))['kaoqin_path'][0])
        # 图片
        self.images = [
            ttk.PhotoImage(file=IMG_PATH / 'icons8_add_folder_24px.png'),
            ttk.PhotoImage(file=IMG_PATH / 'icons8_folder_24px.png'),
            ttk.PhotoImage(file=IMG_PATH / 'icons8_add_folder_24px.png'),
        ]
        # 初始化主题
        setting_df = pd.read_csv(Path.joinpath(SAVE_PATH, 'setting.csv'))
        theme_name = setting_df['current_theme'][0]
        style = ttk.Style()
        style.theme_use(theme_name)

        self.create_paget()
        self.create_pageb()

    def create_paget(self):
        '''1 paget'''
        self.framet = ttk.LabelFrame(self, text='现场人数', bootstyle=SUCCESS)
        self.framet.pack(side=TOP, fill=BOTH, expand=False, padx=10, pady=10)
        # 1.1 选择考勤表
        ttk.Button(
            self.framet,
            text='选择考勤表',
            bootstyle=('outline', WARNING),
            width=9,
            command=self.select_file,
            image=self.images[0],
            compound=LEFT
        ).pack(side=TOP, fill=X, expand=TRUE, padx=20, pady=10)
        mid_frame = ttk.Frame(self.framet)
        mid_frame.pack(side=TOP, fill=BOTH, expand=True)
        for index in [0, 2, 4]:
            mid_frame.columnconfigure(index=index, weight=0)
        for index in [1, 3, 5]:
            mid_frame.columnconfigure(index=index, weight=1)

        self.option_menu_list_month = [i + 1 for i in range(12)]
        self.option_menu_list_month.insert(0, "")
        self.Month_Var = tk.StringVar(value=self.option_menu_list_month[0])
        # 1.2 月份
        ttk.Label(mid_frame, text="月份:").grid(
            row=0,
            column=0,
            sticky="ew",
            padx=(20, 5), pady=(0, 5),
            columnspan=1
        )
        self.OpMenu_Month = ttk.OptionMenu(
            mid_frame,
            self.Month_Var,
            *self.option_menu_list_month,
            bootstyle=('outline', SUCCESS),
        )
        self.OpMenu_Month.config(width=10)
        self.OpMenu_Month.grid(
            row=0,
            column=1,
            sticky="ew",
            padx=(0, 10), pady=(0, 5),
            columnspan=1
        )
        # 1.2 起始日期
        ttk.Label(mid_frame, text="起始日期:").grid(
            row=0,
            column=2,
            sticky="ew",
            padx=(0, 5), pady=(0, 5),
            columnspan=1
        )

        day_list = [i + 1 for i in range(31)]
        self.combobox_start = ttk.Combobox(mid_frame, values=day_list, width=10, style=SUCCESS)
        self.combobox_start.current = ''
        self.combobox_start.grid(
            row=0, column=3, sticky="ew", padx=(0, 10), pady=(0, 5), columnspan=1
        )

        # 1.3终止日期
        ttk.Label(mid_frame, text="终止日期:").grid(
            row=0,
            column=4,
            sticky="ew",
            padx=(0, 5), pady=(0, 5),
            columnspan=1
        )

        self.combobox_end = ttk.Combobox(mid_frame, values=day_list, width=10, style=SUCCESS)
        self.combobox_end.current = ''
        self.combobox_end.grid(
            row=0, column=5, sticky="ew", padx=(0, 20), pady=(0, 5), columnspan=1
        )
        # 1.4 生成并导出
        ttk.Button(
            self.framet,
            text='生成并导出',
            bootstyle=('default', SUCCESS),
            width=9,
            padding=(0, 8),
            command=self.generate,
            compound=LEFT
        ).pack(side=TOP, fill=X, expand=TRUE, padx=20, pady=5)
        # 1.#
        ttk.Label(self.framet,
                  text="功能简介：根据羿云导出的考勤表，导出当月各单位每天现场人数。",
                  padding=(20, 0)
                  ).pack(side=TOP, fill=X, expand=TRUE, pady=(5, 10))

    def create_pageb(self):
        '''2 pageb'''
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(side=BOTTOM, fill=BOTH, expand=True, padx=10, pady=(0, 10))
        # 2.1文件预览
        self.frame1 = ttk.Frame(self.notebook)
        self.notebook.add(self.frame1, text="文件预览")
        self.tableValues1 = initTool.Files().get_file_name(path=kaoqinXlData_PATH)
        self.tableColumns1 = ['序号', '文件名']
        A = initTool.Tab1_1(self.frame1, self.tableValues1, self.tableColumns1)
        self.treeview1 = A.treeview
        # 2.2 各单位人数
        self.frame2 = ttk.Frame(self.notebook)
        self.notebook.add(self.frame2, text="各单位人数")
        self.tableValues2 = []
        self.tableColumns2 = []
        for i in range(31):
            self.tableColumns2.append('')
        self.tableColumns2.insert(0, "单位")
        B = initTool.Tab2_2(self.frame2, self.tableValues2, self.tableColumns2)
        self.treeview2 = B.treeview

    def generate(self):
        '''生成各单位人数和各劳务单位人数的表格'''
        self.tableValues2 = initTool.Files().get_file_name(path=kaoqinXlData_PATH)
        if self.tableValues2 == []:
            Messagebox.show_info_success('未检测到待考勤表！', '温馨提示')
        else:
            month_var = self.Month_Var.get()
            start_day = self.combobox_start.get()
            end_day = self.combobox_end.get()
            if end_day.isnumeric() and start_day.isnumeric() and end_day.isnumeric():
                if int(start_day) in range(1, 32) and int(end_day) in range(1, 32) and int(start_day) <= int(end_day):
                    # 清空KQcsv文件夹里的所有文件
                    initTool.Files().del_kaoqin()
                    # 把考勤表转换成csv文件放在KQcsv文件夹中，文件名为kaoqin.csv
                    filename = self.tableValues2[0][1]
                    kq_df = pd.read_excel(
                        Path.joinpath(kaoqinXlData_PATH, str(filename) + '.xlsx')
                    )
                    last_column_index = int(kq_df.columns[-1])
                    if not last_column_index == 31:
                        for i in range(31 - last_column_index):
                            kq_df.insert(last_column_index + i + 7, last_column_index + i + 1, '')
                    kq_df.to_csv(
                        Path.joinpath(kaoqincsvData_PATH, 'kaoqin.csv'), index=False, encoding='utf-8')
                    co_df = pd.read_csv(Path.joinpath(SAVE_PATH, 'co_dict.csv'))
                    # kaoqin_df 由 kaoqin.csv 转换而来
                    kaoqin_df = pd.read_csv(Path.joinpath(kaoqincsvData_PATH, 'kaoqin.csv'))
                    row_index = []  # row_index 是所有单位列表
                    for co_df_column in co_df.columns:  # co_df_column 是一级单位
                        for item in co_df.loc[1:, co_df_column]:
                            if (pd.isnull(item) == False):
                                row_index.append(item)
                    colunm_index = [i for i in range(int(start_day), int(end_day) + 1)]
                    # kaoqin_data_df 横坐标是start_day~end_day,纵坐标是单位
                    kaoqin_data_df = pd.DataFrame(data=0, index=row_index, columns=colunm_index)
                    for kaoqin_df_row in range(kaoqin_df.shape[0]):  # 遍历 kaoqin_df 每一行,kaoqin_df_row 是当前行数索引值
                        kaoqin_df_row_co = list(kaoqin_df.loc[kaoqin_df_row, :])[3]  # kaoqin_df_row_co 是每一行的第3列,即人员所属单位
                        for kaoqin_data_df_co in kaoqin_data_df.index:  # 遍历所有单位名称
                            if (kaoqin_df_row_co == kaoqin_data_df_co):
                                for day_index in range(int(start_day), int(end_day) + 1):
                                    if (kaoqin_df.iloc[kaoqin_df_row, 6 + day_index] == "√"):
                                        kaoqin_data_df.loc[kaoqin_data_df_co, day_index] += 1
                    kaoqin_final_df = pd.DataFrame(data=0, index=[i for i in co_df.columns],
                                                   columns=[i for i in range(int(start_day), int(end_day) + 1)])
                    # 把分包单位的人数统计在一级单位人数中
                    for kaoqin_data_df_row in list(kaoqin_data_df.index):  # kaoqin_data_df_row 是遍历所有单位
                        for co in list(kaoqin_final_df.index):  # co 是遍历所有一级单位
                            if (kaoqin_data_df_row in list(co_df.loc[:, co])):
                                for i in range(int(start_day), int(end_day) + 1):
                                    kaoqin_final_df.loc[co, i] += list(kaoqin_data_df.loc[kaoqin_data_df_row, :])[
                                        i - int(start_day)]
                    for column in range(kaoqin_final_df.shape[1]):
                        sum_number = 0
                        for row in range(kaoqin_final_df.shape[0]):
                            sum_number += kaoqin_final_df.iloc[row, column]
                    kaoqin_final_df.to_csv(Path.joinpath(kaoqincsvData_PATH, 'primarynumber.csv'))
                    kaoqin_data_df.to_csv(Path.joinpath(kaoqincsvData_PATH, 'allnumber.csv'))
                    # 制作excel表格

                    prj_sim_name = pd.read_csv(Path.joinpath(SAVE_PATH, 'setting.csv'))['simplified_name'][0]
                    if prj_sim_name == ' ':
                        prj_sim_name = ''
                    filename_kaoqin = month_var + "月" + start_day + "日~" + month_var + "月" + end_day + "日" + prj_sim_name + "现场人数" + ".xlsx"
                    with pd.ExcelWriter(Path.joinpath(self.KQ_PATH, filename_kaoqin), engine='openpyxl') as writer:
                        sheet1_df = pd.read_csv(Path.joinpath(kaoqincsvData_PATH, 'primarynumber.csv'))
                        sheet1_df.to_excel(writer, "各单位总人数", index=False)
                        sheet2_df = pd.read_csv(Path.joinpath(kaoqincsvData_PATH, 'allnumber.csv'))
                        sheet2_df.to_excel(writer, "各劳务单位人数", index=False)
                    fontStyle = Font(name='宋体', size=11)
                    bold_fontStyle = Font(name='宋体', size=11, bold="bold")
                    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                    bottom=Side(style='thin'))
                    align = Alignment(horizontal='center', vertical='center')

                    workbook = load_workbook(Path.joinpath(self.KQ_PATH, filename_kaoqin))
                    sheet1 = workbook.worksheets[0]
                    sheet1.cell(1, 1).value = "日期"
                    for column in range(2, int(end_day) - int(start_day) + 3):
                        sheet1.cell(1, column).value = str(month_var) + "月" + str(sheet1.cell(1, column).value) + "日"

                    sheet1.cell(sheet1.max_row + 1, 1).value = "总计"
                    for column in range(2, sheet1.max_column + 1):
                        total = 0
                        for row in range(2, sheet1.max_row):
                            total += sheet1.cell(row, column).value
                        sheet1.cell(sheet1.max_row, column).value = total
                    for row in range(1, sheet1.max_row + 1):
                        for column in range(1, sheet1.max_column + 1):
                            sheet1.cell(row=row, column=column).font = fontStyle
                            sheet1.cell(row=row, column=column).border = border
                            sheet1.cell(row=row, column=column).alignment = align
                    for row in [1, sheet1.max_row]:
                        for column in range(1, sheet1.max_column + 1):
                            sheet1.cell(row=row, column=column).font = bold_fontStyle
                    # 自动调整列宽
                    for col in sheet1.iter_cols():
                        maxLength = 0
                        column = col[0].column_letter
                        for row in range(1, sheet1.max_row + 1):
                            try:
                                if len(str(col[row - 1].value)) > maxLength:
                                    maxLength = len(col[row - 1].value)
                            except:
                                pass
                        width = (maxLength + 2) * 1.8
                        sheet1.column_dimensions[column].width = width
                    sheet2 = workbook.worksheets[1]
                    sheet2.insert_cols(1, 1)
                    for co in co_df.columns:
                        for row in range(1, sheet2.max_row + 1):
                            if co == sheet2.cell(row, 2).value:
                                sheet2.cell(row + 1, 1).value = co
                                sheet2.cell(row, 2).value = "小计"
                    sheet2.delete_rows(2, 1)
                    sheet2.cell(sheet2.max_row + 1, 2).value = "小计"

                    max_row = sheet2.max_row
                    for column in range(3, int(end_day) - int(start_day) + 4):
                        data_list = []
                        total_list = []
                        for row in range(2, max_row + 1):
                            if str(sheet2.cell(row, 2).value) != "小计":
                                data_list.append(sheet2.cell(row, column).value)
                            else:
                                sheet2.cell(row, column).value = int(sum(data_list))
                                total_list.append(int(sum(data_list)))
                                data_list = []
                        sheet2.cell(max_row + 1, column).value = sum(total_list)
                    sheet2.cell(1, 1).value = "一级单位"
                    sheet2.cell(1, 2).value = "二级单位"
                    sheet2.cell(sheet2.max_row, 1).value = "总计"
                    for column in range(3, int(end_day) - int(start_day) + 4):
                        sheet2.cell(1, column).value = str(month_var) + "月" + str(sheet2.cell(1, column).value) + "日"

                    for row in range(1, sheet2.max_row + 1):
                        for column in range(1, sheet2.max_column + 1):
                            sheet2.cell(row=row, column=column).font = fontStyle
                            sheet2.cell(row=row, column=column).border = border
                            sheet2.cell(row=row, column=column).alignment = align
                    row_bold_list = [1, sheet2.max_row]
                    for row in range(1, sheet2.max_row + 1):
                        if sheet2.cell(row=row, column=2).value == "小计":
                            row_bold_list.append(row)

                    for row in row_bold_list:
                        for column in range(1, sheet2.max_column + 1):
                            sheet2.cell(row=row, column=column).font = bold_fontStyle
                    # 自动调整列宽
                    for col in sheet2.iter_cols():
                        maxLength = 0
                        column = col[0].column_letter
                        for row in range(1, sheet2.max_row + 1):
                            try:
                                if len(str(col[row - 1].value)) > maxLength:
                                    maxLength = len(col[row - 1].value)
                            except:
                                pass
                        width = (maxLength + 2) * 1.8
                        sheet2.column_dimensions[column].width = width
                    # 合并单元格
                    sheet2.merge_cells(start_row=sheet2.max_row, start_column=1, end_row=sheet2.max_row, end_column=2)
                    start_row = 2
                    end_row = 2
                    for row in range(2, sheet2.max_row + 1):
                        if sheet2.cell(row, 2).value != "小计":
                            end_row += 1
                        else:
                            sheet2.merge_cells(start_row=start_row, start_column=1, end_row=end_row,
                                               end_column=1)
                            start_row = end_row + 1
                            end_row = start_row
                    workbook.save(Path.joinpath(self.KQ_PATH, filename_kaoqin))
                    self.fresh()
                    Messagebox.show_info_success('完成！', '温馨提示')
                else:
                    Messagebox.show_info_success("日期不合法！", "温馨提示")
            else:
                Messagebox.show_info_success("请输入正确格式的日期！", "温馨提示")

    def fresh(self):
        '''刷新页面'''
        for _ in map(self.treeview2.delete, self.treeview2.get_children("")):
            pass
        try:
            self.tableValues2 = initTool.Files().csv_to_treeviewlist(kaoqincsvData_PATH, 'primarynumber.csv')
            row = 1
            for data in self.tableValues2[1:]:
                self.treeview2.insert('', 'end', text=row, value=data, iid=str(row))
                row += 1
        except:
            pass

    def select_file(self):
        '''选择考勤表'''
        functions.select_file_kq(self)
