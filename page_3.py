from pathlib import Path
import os
import sys

from ttkbootstrap.dialogs import Messagebox
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
import pandas as pd

import initTool
import functions

BASE_PATH = os.path.dirname(os.path.realpath(sys.argv[0]))
IMG_PATH = Path(BASE_PATH) / 'data' / 'image'

SAVE_PATH = Path(BASE_PATH) / 'data' / 'save'
tempDATA_PATH = Path(BASE_PATH) / 'data' / 'temp'
renshuXlData_PATH = Path(tempDATA_PATH) / 'RS'
renshucsvData_PATH = Path(tempDATA_PATH) / 'RScsv'
RS_PATH = Path(BASE_PATH)


class Shishirenshu(ttk.Frame):
    '''实时人数'''

    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        self.root = master
        self.root.showframe.destroy()
        self.root.showframe = self
        self.pack(side=TOP, fill=BOTH, expand=True)
        # 创建temp文件夹
        if not os.path.exists(Path.joinpath(SAVE_PATH)):
            os.makedirs(Path.joinpath(tempDATA_PATH))
        if not os.path.exists(Path.joinpath(tempDATA_PATH, 'RS')):
            os.makedirs(Path.joinpath(tempDATA_PATH, 'RS'))
        if not os.path.exists(Path.joinpath(tempDATA_PATH, 'RScsv')):
            os.makedirs(Path.joinpath(tempDATA_PATH, 'RScsv'))
        # 定义修改后的导出位置
        self.RS_PATH = Path(BASE_PATH)
        if str(pd.read_csv(Path.joinpath(SAVE_PATH, 'setting.csv'))['renshu_path'][0]) != str(
                self.RS_PATH):
            self.RS_PATH = Path(pd.read_csv(Path.joinpath(SAVE_PATH, 'setting.csv'))['renshu_path'][0])
        # 图片
        self.images = [
            ttk.PhotoImage(file=IMG_PATH / 'icons8_add_folder_24px.png'),
            ttk.PhotoImage(file=IMG_PATH / 'icons8_folder_24px.png'),
            ttk.PhotoImage(file=IMG_PATH / 'icons8_add_folder_24px.png'),
        ]
        # 设置主题
        setting_df = pd.read_csv(Path.joinpath(SAVE_PATH, 'setting.csv'))
        theme_name = setting_df['current_theme'][0]
        style = ttk.Style()
        style.theme_use(theme_name)

        self.create_paget()
        self.create_pageb()

    def create_paget(self):
        '''1 paget'''
        self.framet = ttk.LabelFrame(self, text='实时人数', bootstyle=SUCCESS)
        self.framet.pack(side=TOP, fill=BOTH, expand=False, padx=10, pady=10)
        # 1.1 选择现场工人名单
        ttk.Button(
            self.framet,
            text='选择现场工人名单',
            bootstyle=('outline', WARNING),
            command=self.select_files,
            image=self.images[0],
            compound=LEFT
        ).pack(side=TOP, fill=BOTH, expand=True, padx=20, pady=10)
        # 1.2 生成并导出
        ttk.Button(
            self.framet,
            text='生成并导出',
            bootstyle=('default', SUCCESS),
            padding=(0, 8),
            command=self.generate,
            compound=LEFT
        ).pack(side=TOP, fill=BOTH, expand=True, padx=20)
        # 1.#
        ttk.Label(self.framet,
                  text='功能简介：根据羿云导出的”现场工人(在场).xls“，按单位统计各单位实时现场人数。'
                  ).pack(side=TOP, fill=BOTH, expand=True, padx=20, pady=10)

    def create_pageb(self):
        '''2 pageb'''
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(side=BOTTOM, fill=BOTH, expand=True, padx=10, pady=(0, 10))
        # 2.1 文件预览
        self.frame1 = ttk.Frame(self.notebook)
        self.notebook.add(self.frame1, text="文件预览")
        self.tableValues1 = initTool.Files().get_file_name(path=renshuXlData_PATH)
        self.tableColumns1 = ['序号', '文件名']
        A = initTool.Tab1_1(self.frame1, self.tableValues1, self.tableColumns1)
        self.treeview1 = A.treeview
        # 2.2 现场人数预览
        self.frame2 = ttk.Frame(self.notebook)
        self.notebook.add(self.frame2, text="现场人数预览")
        self.tableValues2 = []
        self.tableColumns2 = ['一级单位', '工人数']
        B = initTool.Tab3_2(self.frame2, self.tableValues2, self.tableColumns2)
        self.treeview2 = B.treeview

    def select_files(self):
        '''选择文件把文件保存到RS文件夹中'''
        functions.select_file_rs(self)

    def generate(self):
        '''生成并导出'''
        self.tableValues2 = initTool.Files().get_file_name(path=renshuXlData_PATH)
        if self.tableValues2 == []:
            Messagebox.show_info_success('未检测到表格！', '温馨提示')
        else:
            # 清空之前的数据
            initTool.Files().del_renshu()
            # 把表转成csv文件放在RScsv文件夹中，文件名为renshu.csv
            filename = self.tableValues2[0][1]
            pd.read_excel(
                Path.joinpath(renshuXlData_PATH, str(filename) + '.xlsx')
            ).to_csv(
                Path.joinpath(renshucsvData_PATH, 'renshu.csv'), index=False, encoding='utf-8')
            # 根据renshu.csv,统计各单位人数，文件名为renshu_data.csv
            co_df = pd.read_csv(Path.joinpath(SAVE_PATH, 'co_dict.csv'))
            row_index = []
            for co_df_column in co_df.columns:
                for item in co_df.loc[1:, co_df_column]:
                    if (pd.isnull(item) == False):
                        row_index.append(item)
            # renshu_data_df 用累加法统计各个单位总人数
            renshu_data_df = pd.DataFrame(data=0, index=row_index, columns=[1])
            # renshu_df 用于来源于renshu.csv
            renshu_df = pd.read_csv(Path.joinpath(renshucsvData_PATH, 'renshu.csv'))
            for renshu_df_row in range(renshu_df.shape[0]):
                renshu_df_row_co = list(renshu_df.loc[renshu_df_row, :])[2]
                for renshu_data_df_co in renshu_data_df.index:
                    if (renshu_df_row_co == renshu_data_df_co):
                        renshu_data_df.loc[renshu_data_df_co, 1] += 1
            # 初始化 renshu_final_df
            renshu_final_df = pd.DataFrame(data=0, index=[i for i in co_df.columns], columns=[1])
            # 把子单位的人数归并到对应的一级单位
            for renshu_data_df_row in list(renshu_data_df.index):  # 子单位名称
                for co in list(renshu_final_df.index):
                    if (renshu_data_df_row in list(co_df.loc[:, co])):  # 一级单位名称
                        renshu_final_df.loc[co, 1] += list(renshu_data_df.loc[renshu_data_df_row, :])[0]
            renshu_final_df.columns = ['工人数']
            renshu_final_df.to_csv(Path.joinpath(renshucsvData_PATH, "renshu_final.csv"))
            renshu_final_df.to_excel(Path.joinpath(self.RS_PATH, "实时工人数.xlsx"))
            # 刷新页面
            self.fresh()
            # 美化excel表格
            fontStyle = Font(name='宋体', size=11)
            bold_fontStyle = Font(name='宋体', size=11, bold="bold")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                            bottom=Side(style='thin'))
            align = Alignment(horizontal='center', vertical='center')
            workbook = load_workbook(Path.joinpath(self.RS_PATH, "实时工人数.xlsx"))
            to_be_formated_Xlsheet = workbook.worksheets[0]
            total_row = to_be_formated_Xlsheet.max_row
            to_be_formated_Xlsheet.cell(1, 1).value = "单位"
            to_be_formated_Xlsheet.cell(total_row + 1, 1).value = "总计"

            for row in range(2, total_row + 1):
                for column in [1, 2]:
                    to_be_formated_Xlsheet.cell(row=row, column=column).font = fontStyle
                    to_be_formated_Xlsheet.cell(row=row, column=column).border = border
                    to_be_formated_Xlsheet.cell(row=row, column=column).alignment = align
            for row in [1, total_row + 1]:
                for column in [1, 2]:
                    to_be_formated_Xlsheet.cell(row=row, column=column).font = bold_fontStyle
                    to_be_formated_Xlsheet.cell(row=row, column=column).border = border
                    to_be_formated_Xlsheet.cell(row=row, column=column).alignment = align
            to_be_formated_Xlsheet.cell(row=total_row + 1, column=2).value = sum(
                [to_be_formated_Xlsheet.cell(row=i, column=2).value for i in range(2, total_row + 1)])
            # 自动调整列宽
            for col in to_be_formated_Xlsheet.iter_cols():
                maxLength = 0
                column = col[0].column_letter
                for row in range(1, to_be_formated_Xlsheet.max_row + 1):
                    try:
                        if len(str(col[row - 1].value)) > maxLength:
                            maxLength = len(col[row - 1].value)
                    except:
                        pass
                width = (maxLength + 2) * 1.8
                to_be_formated_Xlsheet.column_dimensions[column].width = width
            workbook.save(Path.joinpath(self.RS_PATH, "实时工人数.xlsx"))
            Messagebox.show_info_success('完成！', '温馨提示')

    def fresh(self):
        '''刷新页面'''
        for _ in map(self.treeview2.delete, self.treeview2.get_children("")):
            pass
        try:
            self.tableValues2 = initTool.Files().csv_to_treeviewlist(renshucsvData_PATH, 'renshu_final.csv')
            row = 1
            for data in self.tableValues2[1:]:
                self.treeview2.insert('', 'end', text=row, value=data, iid=str(row))
                row += 1
        except:
            pass
