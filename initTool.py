import os
import sys

import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import tkinter as tk
import openpyxl
from pathlib import Path
import csv
import datetime
import pandas as pd
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl import load_workbook
import pyperclip

BASE_PATH = os.path.dirname(os.path.realpath(sys.argv[0]))
tempDATA_PATH = Path(BASE_PATH) / 'data' / 'temp'

XlFilesToBeMerged_PATH = Path(tempDATA_PATH) / 'XlFiles'
csvFilesToBeMerged_PATH = Path(tempDATA_PATH) / 'csvFiles'
appendXlData_PATH = Path(tempDATA_PATH) / 'AD'
appendcsvData_PATH = Path(tempDATA_PATH) / 'ADcsv'

renshuXlData = Path(tempDATA_PATH) / 'RS'
SAVE_PATH = Path(BASE_PATH) / 'data' / 'save'


class Tab1_1:
    '''
    颜色：绿色
    x轴滚动条：无
    y轴滚动条：有
    列数：2
    是否需要复制名字：不需要
    '''

    def __init__(self, frame, tableValues, tableColumns):
        # 滚动条
        self.treeScroly = tk.Scrollbar(frame)
        self.treeScroly.pack(side="right", fill="y")
        self.treeview = ttk.Treeview(frame, selectmode="extended", columns=tableColumns,
                                     yscrollcommand=self.treeScroly.set,
                                     show='headings',
                                     height=40, bootstyle=SUCCESS)

        self.treeview.pack(expand=TRUE, fill=BOTH)
        self.treeScroly.config(command=self.treeview.yview)

        for i in range(len(tableColumns)):
            self.treeview.heading(column=tableColumns[i], text=tableColumns[i], anchor=CENTER)
            self.treeview.column(tableColumns[i], minwidth=100, anchor=CENTER, stretch=True)
        row = 1
        for data in tableValues:
            self.treeview.insert('', 'end', text=row, value=data, iid=row)
            row += 1
        self.treeview.column('#1', width=40, minwidth=40)
        self.treeview.column('#2', width=90, minwidth=40)


class Tab1_2:
    '''
    颜色：绿色
    x轴滚动条：有
    y轴滚动条：有
    列数：11
    是否需要复制名字：需要
    '''

    def __init__(self, frame, tableValues, tableColumns):
        self.treeScroly = tk.Scrollbar(frame)
        self.treeScroly.pack(side="right", fill="y")
        self.treeScrolx = tk.Scrollbar(frame, orient=HORIZONTAL)
        self.treeScrolx.pack(side="bottom", fill="x")
        self.treeview = ttk.Treeview(frame, selectmode="extended", columns=tableColumns,
                                     yscrollcommand=self.treeScroly.set,
                                     xscrollcommand=self.treeScrolx.set, show='headings',
                                     height=40, bootstyle=SUCCESS)

        self.treeview.pack(expand=TRUE, fill=BOTH)
        self.treeScroly.config(command=self.treeview.yview)
        self.treeScrolx.config(command=self.treeview.xview)

        for i in range(len(tableColumns)):
            self.treeview.heading(column=tableColumns[i], text=tableColumns[i], anchor=CENTER)
            self.treeview.column(tableColumns[i], minwidth=100, anchor=CENTER, stretch=True)
        row = 1
        for data in tableValues:
            self.treeview.insert('', 'end', text=row, value=data, iid=row)
            row += 1
        self.treeview.column('#1', width=40, minwidth=10)  # 序号
        self.treeview.column('#2', width=50, minwidth=10)  # 单位
        self.treeview.column('#3', width=50, minwidth=0)  # 姓名
        self.treeview.column('#4', width=55, minwidth=5)  # 工种
        self.treeview.column('#5', width=40, minwidth=4)  # 性别
        self.treeview.column('#6', width=40, minwidth=4)  # 年龄
        self.treeview.column('#7', width=65, minwidth=14)  # 身份证号
        self.treeview.column('#8', width=65, minwidth=10)  # 地址
        self.treeview.column('#9', width=65, minwidth=10)  # 电话号码
        self.treeview.column('#10', width=74, minwidth=10)  # 培训时间
        self.treeview.column('#11', width=70, minwidth=10)  # 备注
        self.treeview.bind("<Control-Key-C>", lambda x: self.copy_from_treeview(x))
        self.treeview.bind("<Control-Key-c>", lambda x: self.copy_from_treeview(x))

    def copy_from_treeview(self, event):
        '''复制控件'''
        selection = self.treeview.selection()
        copy_values = []
        for each in selection:
            try:
                value = self.treeview.item(each)["values"]
                copy_values.append(list(value)[2])
            except:
                pass
        copy_string = "\n".join(copy_values)
        pyperclip.copy(copy_string)


class Tab2_2:
    '''
     颜色：绿色
     x轴滚动条：有
     y轴滚动条：有
     列数：32
     是否需要复制名字：不需要
     '''

    def __init__(self, frame, tableValues, tableColumns):

        self.treeScroly = tk.Scrollbar(frame)
        self.treeScroly.pack(side="right", fill="y")
        self.treeScrolx = tk.Scrollbar(frame, orient=HORIZONTAL)
        self.treeScrolx.pack(side="bottom", fill="x")
        self.treeview = ttk.Treeview(frame, selectmode="extended", columns=tableColumns,
                                     yscrollcommand=self.treeScroly.set,
                                     xscrollcommand=self.treeScrolx.set, show='headings',
                                     height=40, bootstyle=SUCCESS)

        self.treeview.pack(expand=TRUE, fill=BOTH)
        self.treeScroly.config(command=self.treeview.yview)
        self.treeScrolx.config(command=self.treeview.xview)

        for i in range(len(tableColumns)):
            self.treeview.heading(column=tableColumns[i], text=tableColumns[i], anchor=CENTER)
            self.treeview.column(tableColumns[i], minwidth=100, anchor=CENTER, stretch=True)
        row = 1
        for data in tableValues:
            self.treeview.insert('', 'end', text=row, value=data, iid=row)
            row += 1

        self.treeview.column('#1', width=100, minwidth=100)
        for i in range(31):
            var = '#' + str(i)
            self.treeview.column(var, width=50, minwidth=10)


class Tab3_2:
    '''
    颜色：绿色
    x轴滚动条：有
    y轴滚动条：有
    列数：2
    是否需要复制名字：不需要
    '''

    def __init__(self, frame, tableValues, tableColumns):

        # Scrollbar
        self.treeScroly = tk.Scrollbar(frame)
        self.treeScroly.pack(side="right", fill="y")
        self.treeScrolx = tk.Scrollbar(frame, orient=HORIZONTAL)
        self.treeScrolx.pack(side="bottom", fill="x")
        self.treeview = ttk.Treeview(frame, selectmode="extended", columns=tableColumns,
                                     yscrollcommand=self.treeScroly.set,
                                     xscrollcommand=self.treeScrolx.set, show='headings',
                                     height=40, bootstyle=SUCCESS)

        self.treeview.pack(expand=TRUE, fill=BOTH)
        self.treeScroly.config(command=self.treeview.yview)
        self.treeScrolx.config(command=self.treeview.xview)

        for i in range(len(tableColumns)):
            self.treeview.heading(column=tableColumns[i], text=tableColumns[i], anchor=CENTER)
            self.treeview.column(tableColumns[i], minwidth=100, anchor=CENTER, stretch=True)
        row = 1
        for data in tableValues:
            self.treeview.insert('', 'end', text=row, value=data, iid=row)
            row += 1
        self.treeview.column('#1', width=100, minwidth=40)  # 一级单位
        self.treeview.column('#2', width=100, minwidth=40)  # 工人数


class Files:
    '''处理文件用类'''

    def file_to_csv(self, sheet, filepath, filename):
        '''将sheet转换成csv文件'''
        with open(Path.joinpath(filepath, filename), 'w', encoding='utf-8', newline='') as csvfile:
            writer = csv.writer(csvfile)
            for row in iter(sheet):
                if row[2].value != None:
                    rowlist = [cell.value for cell in row]
                    # 去除空格
                    for i in range(len(rowlist)):
                        rowlist[i] = str(rowlist[i]).replace(" ", "")
                        rowlist[i] = str(rowlist[i]).replace("\n", "")
                        # 去除None
                        if rowlist[i] == 'None':
                            rowlist[i] = ''
                    # 计算年龄
                    if len(rowlist[6]) == 18:
                        birth_year = int(rowlist[6][6:10])
                        birth_month = int(rowlist[6][10:12])
                        birth_day = int(rowlist[6][12:14])
                        today = datetime.date.today()
                        age = today.year - birth_year - ((today.month, today.day) < (birth_month, birth_day))
                        if age < 0 or age > 100:
                            pass
                        else:
                            rowlist[5] = age
                    else:
                        try:
                            birth_year = int(rowlist[6][6:10])
                            birth_month = int(rowlist[6][10:12])
                            birth_day = int(rowlist[6][12:14])
                            today = datetime.date.today()
                            age = today.year - birth_year - ((today.month, today.day) < (birth_month, birth_day))
                            if age < 0 or age > 100:
                                birth_year = int(rowlist[6][5:9])
                                birth_month = int(rowlist[6][9:11])
                                birth_day = int(rowlist[6][11:13])
                                today = datetime.date.today()
                                age = today.year - birth_year - ((today.month, today.day) < (birth_month, birth_day))
                                rowlist[5] = age
                            else:
                                rowlist[5] = age
                        except:
                            pass
                    if len(rowlist[4]) != 1:
                        if rowlist[6][-2].isdigit() == True:
                            if int(rowlist[6][-2]) % 2 == 0:
                                rowlist[4] = "女"
                            else:
                                rowlist[4] = "男"
                    writer.writerow(rowlist)
            csvfile.close()

    def file_to_csv2(self, sheet, filepath, filename):
        # sheet将会转换成csv文件，filepath，filename分别为csv的路径和文件名
        # 用来将11列的excel文件转换成csv，表头列名来自excel的第三行，第二行的第二列应为空值
        with open(Path.joinpath(filepath, filename), 'w', encoding='utf-8', newline='') as csvfile:
            writer = csv.writer(csvfile)
            for row in iter(sheet):
                if row[2].value != None and row[2].value != '':
                    if row[2].value == "姓名":
                        # 初始化每个子csv的第一行：序号,单位,姓名,工种,性别,年龄,身份证号,家庭住址,手机,退场时间,备注,
                        rowlist = ['序号', '单位', '姓名', '工种', '性别', '年龄', '身份证号', '家庭住址', '手机', '退场时间', '备注']
                    else:
                        #
                        rowlist = [cell.value for cell in row]
                        rowlist = [rowlist[i] for i in range(11)]

                    # 去除空格
                    for i in range(len(rowlist)):
                        rowlist[i] = str(rowlist[i]).replace(" ", "")
                        rowlist[i] = str(rowlist[i]).replace("\n", "")
                        # 去除None
                        if rowlist[i] == 'None':
                            rowlist[i] = ''
                    # 计算年龄
                    if len(rowlist[6]) == 18:
                        birth_year = int(rowlist[6][6:10])
                        birth_month = int(rowlist[6][10:12])
                        birth_day = int(rowlist[6][12:14])
                        today = datetime.date.today()
                        age = today.year - birth_year - ((today.month, today.day) < (birth_month, birth_day))
                        if age < 0 or age > 100:
                            pass
                        else:
                            rowlist[5] = age
                    else:
                        try:
                            birth_year = int(rowlist[6][6:10])
                            birth_month = int(rowlist[6][10:12])
                            birth_day = int(rowlist[6][12:14])
                            today = datetime.date.today()
                            age = today.year - birth_year - ((today.month, today.day) < (birth_month, birth_day))
                            if age < 0 or age > 100:
                                birth_year = int(rowlist[6][5:9])
                                birth_month = int(rowlist[6][9:11])
                                birth_day = int(rowlist[6][11:13])
                                today = datetime.date.today()
                                age = today.year - birth_year - ((today.month, today.day) < (birth_month, birth_day))
                                rowlist[5] = age
                            else:
                                rowlist[5] = age
                        except:
                            pass
                    if len(rowlist[4]) != 1:
                        if rowlist[6][-2].isdigit() == True:
                            if int(rowlist[6][-2]) % 2 == 0:
                                rowlist[4] = "女"
                            else:
                                rowlist[4] = "男"
                    writer.writerow(rowlist)
            csvfile.close()

    def get_file_name(self, path):
        '''获取目录下以”.xlsx“或".xls"结尾的文件，并返回一个列表'''
        filelist = []
        num = 1
        for entry in Path(path).iterdir():
            if str(entry).endswith('.xlsx') or str(entry).endswith('.xls'):
                listtemp = []
                listtemp.append(num)
                listtemp.append(entry.stem)
                filelist.append(listtemp)
                num += 1
        return filelist

    def del_kaoqin(self):
        '''删除temp->KQcsv里的文件'''
        for entry in (Path(tempDATA_PATH) / 'KQcsv').iterdir():
            if entry.is_file() or entry.is_symlink():
                entry.unlink()

    def del_renshu(self):
        '''删除temp->RScsv里的文件'''
        for entry in (Path(tempDATA_PATH) / 'RScsv').iterdir():
            if entry.is_file() or entry.is_symlink():
                entry.unlink()

    def del_temp_csv_ADcsv(self):
        '''删除ADcsv文件夹里的所有文件'''
        for entry in appendcsvData_PATH.iterdir():
            if entry.is_file() or entry.is_symlink():
                entry.unlink()

    def csv_to_treeviewlist(self, filepath, filename):
        '''把csv文件转换成treeview文件'''
        with open(filepath / filename, 'r', encoding='utf-8') as csvfile:
            data = csvfile.readlines()
            treeviewlist = []
            for item in data:
                line = item[:-1]
                rowlist = line.split(',')
                try:
                    rowlist[8] = rowlist[8][:11]
                except:
                    pass
                treeviewlist.append(rowlist)
            csvfile.close()
            return treeviewlist

    def merge_csv_files(self):
        '''合并csv文件'''
        dataList = []
        try:
            for csvfile in Path(csvFilesToBeMerged_PATH).iterdir():
                if str(csvfile).endswith('.csv') and csvfile.stem != 'total':
                    data = pd.read_csv(csvfile)
                    dataList.append(data)
            allData = pd.concat(dataList, axis=0)
            allData.to_csv(Path.joinpath(csvFilesToBeMerged_PATH, 'total.csv'), index=False)
        except:
            pass

    def fresh_csv(self, filepath, filename):
        '''刷新csv文件的第一列序号'''
        try:
            df = pd.read_csv(Path.joinpath(filepath, filename))
            df['序号'] = [i for i in range(1, df.shape[0] + 1)]
            df.to_csv(Path.joinpath(filepath, filename), index=False)
        except:
            pass

    def csv_to_excel(self, today, filepath, filename):
        '''
        创建一个excel文件，保存total.csv的数据
        today: 培训时间
        filepath: 保存的路径
        filename: 保存的文件名
        '''
        for excelfile in filepath.iterdir():
            if str(excelfile).endswith('.xlsx') and excelfile.name == filename:
                wb = openpyxl.load_workbook(Path.joinpath(filepath, filename))
                sheet_names = wb.get_sheet_names()
                for sheet in sheet_names:
                    wb.remove_sheet(wb[sheet])
                wb.create_sheet()
                wb.save(Path.joinpath(filepath, filename))
                wb = load_workbook(Path.joinpath(filepath, filename))
                ws = wb.worksheets[0]

                datalist = self.csv_to_treeviewlist(csvFilesToBeMerged_PATH, 'total.csv')

                for row in range(4, len(datalist) + 4):
                    ws.cell(row=row, column=7).number_format = '@'
                    ws.cell(row=row, column=9).number_format = '@'
                fontStyle = Font(name='宋体', size=11)
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                bottom=Side(style='thin'))
                align = Alignment(horizontal='center', vertical='center')
                ws.insert_rows(1, 2)
                ws.cell(row=1, column=1).value = '人员进场汇总表'
                prj_name = pd.read_csv(Path.joinpath(SAVE_PATH, 'setting.csv'))['prj_name'][0]
                ws.cell(row=2, column=1).value = '工程名称：' + str(prj_name)
                ws.cell(row=2, column=10).value = '培训日期：'
                ws.cell(row=2, column=11).value = datetime.datetime.strftime(today, '%Y.%m.%d')
                # 写入数据
                for row in range(1, len(datalist) + 1):
                    for col in range(1, len(datalist[row - 1]) + 1):
                        ws.cell(row=row + 2, column=col).value = datalist[row - 1][col - 1]
                # 格式
                for row in range(3, len(datalist) + 3):
                    for column in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=column).font = fontStyle
                        ws.cell(row=row, column=column).border = border
                        ws.cell(row=row, column=column).alignment = align
                # 自动调整列宽
                for col in ws.iter_cols():
                    maxLength = 0
                    column = col[0].column_letter
                    for row in range(3, ws.max_row + 1):
                        try:
                            if len(str(col[row - 1].value)) > maxLength:
                                maxLength = len(col[row - 1].value)
                        except:
                            pass
                    width = (maxLength + 2) * 1.8
                    ws.column_dimensions[column].width = width
                # 合并单元格
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
                ws.cell(row=1, column=1).font = Font(name='黑体', size=18, bold=True)
                ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
                wb.save(Path.joinpath(filepath, filename))
                wb.close()

            else:
                wb = openpyxl.Workbook(Path.joinpath(filepath, filename))
                wb.create_sheet()
                wb.save(Path.joinpath(filepath, filename))
                wb = load_workbook(Path.joinpath(filepath, filename))
                ws = wb.worksheets[0]

                datalist = self.csv_to_treeviewlist(csvFilesToBeMerged_PATH, 'total.csv')

                for row in range(4, len(datalist) + 4):
                    ws.cell(row=row, column=7).number_format = '@'
                    ws.cell(row=row, column=9).number_format = '@'
                fontStyle = Font(name='宋体', size=11)
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                bottom=Side(style='thin'))
                align = Alignment(horizontal='center', vertical='center')
                ws.insert_rows(1, 2)
                ws.cell(row=1, column=1).value = '人员进场汇总表'
                prj_name = pd.read_csv(Path.joinpath(SAVE_PATH, 'setting.csv'))['prj_name'][0]
                ws.cell(row=2, column=1).value = '工程名称：' + str(prj_name)
                ws.cell(row=2, column=10).value = '培训日期：'
                ws.cell(row=2, column=11).value = datetime.datetime.strftime(today, '%Y.%m.%d')
                # 写入数据
                for row in range(1, len(datalist) + 1):
                    for col in range(1, len(datalist[row - 1]) + 1):
                        ws.cell(row=row + 2, column=col).value = datalist[row - 1][col - 1]
                # 格式
                for row in range(3, len(datalist) + 3):
                    for column in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=column).font = fontStyle
                        ws.cell(row=row, column=column).border = border
                        ws.cell(row=row, column=column).alignment = align
                # 自动调整列宽
                for col in ws.iter_cols():
                    maxLength = 0
                    column = col[0].column_letter
                    for row in range(3, ws.max_row + 1):
                        try:
                            if len(str(col[row - 1].value)) > maxLength:
                                maxLength = len(col[row - 1].value)
                        except:
                            pass
                    width = (maxLength + 2) * 1.8
                    ws.column_dimensions[column].width = width
                # 合并单元格
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
                ws.cell(row=1, column=1).font = Font(name='黑体', size=18, bold=True)
                ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
                wb.save(Path.joinpath(filepath, filename))
                wb.close()

    def csv_to_excel2(self, csvpath, csvname, filepath, filename):
        '''
        用于导出名单总表，和每天的汇总表区别在于不显示培训时间
        '''
        wb = openpyxl.Workbook(Path.joinpath(filepath, filename))
        wb.create_sheet()
        wb.save(Path.joinpath(filepath, filename))
        wb = load_workbook(Path.joinpath(filepath, filename))
        ws = wb.worksheets[0]

        datalist = self.csv_to_treeviewlist(csvpath, csvname)

        for row in range(4, len(datalist) + 4):
            ws.cell(row=row, column=7).number_format = '@'
            ws.cell(row=row, column=9).number_format = '@'
        fontStyle = Font(name='宋体', size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))
        align = Alignment(horizontal='center', vertical='center')
        ws.insert_rows(1, 2)
        ws.cell(row=1, column=1).value = '人员进场汇总表'
        prj_name = pd.read_csv(Path.joinpath(SAVE_PATH, 'setting.csv'))['prj_name'][0]
        ws.cell(row=2, column=1).value = '工程名称：' + str(prj_name)
        # 写入数据
        for row in range(1, len(datalist) + 1):
            for col in range(1, len(datalist[row - 1]) + 1):
                ws.cell(row=row + 2, column=col).value = datalist[row - 1][col - 1]
        # 格式
        for row in range(3, len(datalist) + 3):
            for column in range(1, ws.max_column + 1):
                ws.cell(row=row, column=column).font = fontStyle
                ws.cell(row=row, column=column).border = border
                ws.cell(row=row, column=column).alignment = align
        # 自动调整列宽
        for col in ws.iter_cols():
            maxLength = 0
            column = col[0].column_letter
            for row in range(3, ws.max_row + 1):
                try:
                    if len(str(col[row - 1].value)) > maxLength:
                        maxLength = len(col[row - 1].value)
                except:
                    pass
            width = (maxLength + 2) * 1.8
            ws.column_dimensions[column].width = width
        # 合并单元格
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        ws.cell(row=1, column=1).font = Font(name='黑体', size=18, bold=True)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        wb.save(Path.joinpath(filepath, filename))
        wb.close()


if __name__ == '__main__':
    tableValues1 = Files().get_file_name(path=XlFilesToBeMerged_PATH)
    for i in tableValues1:
        myworksheet = load_workbook(Path.joinpath(XlFilesToBeMerged_PATH, str(i[1]) + '.xlsx')).worksheets[0]
        Files().file_to_csv(myworksheet, tempDATA_PATH, 'file' + str(i[0]) + '.csv')
    Files().merge_csv_files()
    today = datetime.date.today()
    year = today.year
    month = today.month
    day = today.day
